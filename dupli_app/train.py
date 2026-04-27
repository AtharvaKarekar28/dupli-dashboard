import numpy as np
import pandas as pd
import pickle
import openpyxl
from sklearn.ensemble import GradientBoostingRegressor
from sklearn.model_selection import cross_val_score
import warnings
warnings.filterwarnings('ignore')

FEATURES    = ['Qty', 'Setup_min', 'Log_Qty']
MODEL_PATH  = 'gb_before.pkl'
DATA_PATH_1 = 'data/production_system.xlsx'
DATA_PATH_2 = 'data/syracuse_digital.xlsx'

def load_file(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    h  = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    df = pd.DataFrame([dict(zip(h, r)) for r in ws.iter_rows(min_row=2, values_only=True)])
    return df

def process(df):
    df['Start']   = pd.to_datetime(df['Start'] if 'Start' in df.columns else df.get('Start_Date'), errors='coerce')
    df['Stop']    = pd.to_datetime(df['Stop']  if 'Stop'  in df.columns else df.get('Stop_Date'),  errors='coerce')
    df            = df[df['Start'].notna() & df['Stop'].notna()].copy()
    df['Machine'] = df['Dept'].apply(lambda d: 'Memjet_1' if 'Memjet 1' in str(d) else 'Other')
    df['Elapsed'] = pd.to_numeric(df.get('Elapsed', 0), errors='coerce').fillna(0)
    df['Prod_Q']  = pd.to_numeric(df.get('Prod Q.', 0), errors='coerce').fillna(0)
    return df

def train():
    print("Loading both production files...")
    df1 = process(load_file(DATA_PATH_1))
    df2 = process(load_file(DATA_PATH_2))
    df  = pd.concat([df1, df2], ignore_index=True)
    print(f"Combined records: {len(df):,}")

    df_m1 = df[df['Machine']=='Memjet_1'].sort_values(['Employee','Start']).copy()
    df_m1['Employee']  = df_m1['Employee'].fillna('Unknown')
    df_m1['Prev_Stop'] = df_m1.groupby('Employee')['Stop'].shift(1)
    df_m1['Gap_min']   = (df_m1['Start'] - df_m1['Prev_Stop']).dt.total_seconds() / 60
    df_m1['Gap_min']   = df_m1['Gap_min'].clip(0, 30)
    df_m1['Full_Cycle_min'] = df_m1['Elapsed'] + df_m1['Gap_min']
    df_m1['Log_Qty']        = np.log(df_m1['Prod_Q'].clip(1))

    df_clean = df_m1[
        (df_m1['Full_Cycle_min'] > 0) &
        (df_m1['Full_Cycle_min'] < 120) &
        (df_m1['Prod_Q'] > 0)
    ].rename(columns={'Prod_Q':'Qty', 'Elapsed':'Setup_min'})

    X = df_clean[FEATURES]
    y = df_clean['Full_Cycle_min']
    print(f"Training on {len(df_clean):,} records...")

    model = GradientBoostingRegressor(
        n_estimators=200, max_depth=3,
        learning_rate=0.05, random_state=42
    )
    model.fit(X, y)

    cv_r2  = cross_val_score(model, X, y, cv=5, scoring='r2').mean()
    cv_mae = -cross_val_score(model, X, y, cv=5, scoring='neg_mean_absolute_error').mean()

    with open(MODEL_PATH, 'wb') as f:
        pickle.dump(model, f)

    print(f"\n✓ Model saved to {MODEL_PATH}")
    print(f"   Records : {len(df_clean):,}  |  CV R² : {cv_r2:.3f}  |  CV MAE : {cv_mae:.2f} min")

if __name__ == '__main__':
    train()